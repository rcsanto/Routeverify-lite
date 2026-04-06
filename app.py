import streamlit as st
import pandas as pd
import os
import tempfile
import base64
import io
import zipfile
from datetime import datetime
from dotenv import load_dotenv
import anthropic
from pypdf import PdfReader
import json
import logging
from typing import Dict, List, Optional
import re
import openpyxl
from openpyxl import load_workbook
from copy import copy
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv()

st.set_page_config(page_title="RouteVerify - DSNY", layout="wide")
st.markdown("# 🗑️ RouteVerify — DSNY", unsafe_allow_html=False)

st.markdown("""
<style>
/* ── Google Font ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

/* ── Base ── */
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
}

/* ── Hide Streamlit branding ── */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* ── Main container padding ── */
.block-container {
    padding-top: 1rem !important;
    padding-bottom: 2rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}

/* ── App title ── */
h1 {
    font-size: 1.4rem !important;
    font-weight: 700 !important;
    color: #1a6b2f !important;
    margin-bottom: 0.5rem !important;
}

h2, h3 {
    font-size: 1.1rem !important;
    font-weight: 600 !important;
    color: #1a3a1f !important;
}

/* ── Buttons — larger tap targets ── */
.stButton > button {
    border-radius: 10px !important;
    padding: 0.55rem 1rem !important;
    font-size: 0.88rem !important;
    font-weight: 600 !important;
    min-height: 44px !important;
    width: 100% !important;
    transition: all 0.15s ease !important;
    border: none !important;
}

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1a6b2f, #2d9e4f) !important;
    color: white !important;
}

.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #155826, #27894a) !important;
    box-shadow: 0 4px 12px rgba(26,107,47,0.35) !important;
    transform: translateY(-1px) !important;
}

.stButton > button[kind="secondary"] {
    background: #f5f5f5 !important;
    color: #333 !important;
    border: 1px solid #ddd !important;
}

.stButton > button[kind="secondary"]:hover {
    background: #ffe5e5 !important;
    border-color: #e53935 !important;
    color: #e53935 !important;
}

/* ── Download buttons ── */
.stDownloadButton > button {
    border-radius: 10px !important;
    padding: 0.55rem 1rem !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    min-height: 44px !important;
    width: 100% !important;
    background: linear-gradient(135deg, #1565c0, #1e88e5) !important;
    color: white !important;
    border: none !important;
}

/* ── Cards / containers ── */
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlock"] {
    background: white;
    border-radius: 14px;
    padding: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    margin-bottom: 0.75rem;
}

/* ── Progress bar ── */
.stProgress > div > div > div {
    height: 10px !important;
    border-radius: 5px !important;
    background: linear-gradient(90deg, #1a6b2f, #4caf50) !important;
}

.stProgress > div > div {
    background: #e0e0e0 !important;
    border-radius: 5px !important;
    height: 10px !important;
}

/* ── Text inputs ── */
.stTextInput > div > div > input {
    border-radius: 8px !important;
    border: 1.5px solid #ddd !important;
    padding: 0.5rem 0.75rem !important;
    font-size: 0.9rem !important;
    min-height: 44px !important;
    transition: border-color 0.2s !important;
}

.stTextInput > div > div > input:focus {
    border-color: #1a6b2f !important;
    box-shadow: 0 0 0 2px rgba(26,107,47,0.15) !important;
}

/* ── Text area ── */
.stTextArea > div > div > textarea {
    border-radius: 8px !important;
    border: 1.5px solid #ddd !important;
    font-size: 0.88rem !important;
    transition: border-color 0.2s !important;
}

.stTextArea > div > div > textarea:focus {
    border-color: #1a6b2f !important;
    box-shadow: 0 0 0 2px rgba(26,107,47,0.15) !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #1a6b2f !important;
    border-radius: 12px !important;
    padding: 1rem !important;
    background: #f8fdf9 !important;
}

/* ── Alerts ── */
.stSuccess {
    border-radius: 10px !important;
    font-weight: 500 !important;
}
.stError {
    border-radius: 10px !important;
    font-weight: 500 !important;
}
.stWarning {
    border-radius: 10px !important;
    font-weight: 500 !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #1a3a1f !important;
}

[data-testid="stSidebar"] * {
    color: #e8f5e9 !important;
}

[data-testid="stSidebar"] .stTextInput > div > div > input {
    background: #2d5a35 !important;
    color: white !important;
    border-color: #3d7a45 !important;
}

[data-testid="stSidebar"] .stButton > button {
    background: #2d5a35 !important;
    color: #e8f5e9 !important;
    border: 1px solid #3d7a45 !important;
}

[data-testid="stSidebar"] .stCheckbox label {
    color: #e8f5e9 !important;
}

/* ── Expander ── */
[data-testid="stExpander"] {
    border-radius: 12px !important;
    border: 1px solid #e0e0e0 !important;
    overflow: hidden !important;
}

[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 0.75rem 1rem !important;
    background: #f8fdf9 !important;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden !important;
    font-size: 0.82rem !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px !important;
    background: #f0f0f0 !important;
    border-radius: 10px !important;
    padding: 4px !important;
}

.stTabs [data-baseweb="tab"] {
    border-radius: 8px !important;
    padding: 0.4rem 1rem !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    min-height: 40px !important;
}

.stTabs [aria-selected="true"] {
    background: white !important;
    color: #1a6b2f !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.12) !important;
}

/* ── Divider ── */
hr {
    border-color: #e0e0e0 !important;
    margin: 1rem 0 !important;
}

/* ── Mobile responsive — single column on narrow screens ── */
@media (max-width: 768px) {
    .block-container {
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
    }

    h1 {
        font-size: 1.15rem !important;
    }

    .stButton > button {
        font-size: 0.82rem !important;
        padding: 0.5rem 0.75rem !important;
    }

    /* Stack columns on mobile */
    [data-testid="column"] {
        min-width: 100% !important;
        flex: 1 1 100% !important;
    }
}

/* ── Checkbox — bigger touch target ── */
.stCheckbox label {
    font-size: 0.88rem !important;
    min-height: 36px !important;
    display: flex !important;
    align-items: center !important;
}

/* ── Spinner ── */
.stSpinner {
    color: #1a6b2f !important;
}

/* ── Toast ── */
[data-testid="stToast"] {
    border-radius: 12px !important;
    font-weight: 500 !important;
}

/* ── Select/date inputs ── */
.stDateInput input, .stSelectbox select {
    border-radius: 8px !important;
    min-height: 44px !important;
    font-size: 0.9rem !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="background:linear-gradient(90deg,#1a6b2f,#2d9e4f);padding:0.6rem 1rem;border-radius:10px;margin-bottom:1rem;display:flex;align-items:center;justify-content:space-between;">
  <span style="color:white;font-weight:700;font-size:1rem;">🗑️ RouteVerify Lite</span>
  <span style="color:#c8e6c9;font-size:0.8rem;">DSNY Supervisor Dashboard</span>
</div>
""", unsafe_allow_html=True)

# ─── SIDEBAR ───────────────────────────────────────────────────────────────────

_env_key = os.getenv("CLAUDE_API_KEY", "")
with st.sidebar:
    st.header("Configuration")
    debug_mode = st.checkbox("Debug Mode")
    _api_key = st.text_input("Anthropic API Key", value=_env_key, type="password", help="Paste your sk-ant-... key here")
    st.divider()
    st.subheader("🗑️ Clear All Routes")
    confirm_clear = st.checkbox("Confirm clear all routes")
    if st.button("Clear All Routes", disabled=not confirm_clear):
        st.session_state.routes = []
        st.rerun()

    st.divider()
    st.subheader("🏢 Garage / Command")
    garage_val = st.text_input("Garage", placeholder="e.g. Manhattan 1", key="garage_input")
    if 'garage' not in st.session_state:
        st.session_state.garage = ''
    if garage_val != st.session_state.garage:
        st.session_state.garage = garage_val

    st.divider()
    st.subheader("💾 Session")
    # Save
    if st.session_state.get('routes'):
        save_data = []
        for r in st.session_state.routes:
            entry = {k: v for k, v in r.items() if k not in ('df', 'gps_streets')}
            entry['df'] = r['df'].to_dict(orient='records') if hasattr(r.get('df'), 'to_dict') else []
            save_data.append(entry)
        save_json = json.dumps(save_data, indent=2)
        st.download_button("💾 Save Session", data=save_json,
                           file_name=f"routeverify_session_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                           mime="application/json", key="dl_save_session")
    # Load
    session_file = st.file_uploader("📂 Load Session", type=["json"], key="load_session_file")
    if session_file:
        try:
            loaded = json.loads(session_file.read())
            for entry in loaded:
                if 'df' in entry and isinstance(entry['df'], list):
                    entry['df'] = pd.DataFrame(entry['df'])
                if 'gps_streets' not in entry:
                    entry['gps_streets'] = set()
                if 'manual_overrides' not in entry:
                    entry['manual_overrides'] = {}
            st.session_state.routes = loaded
            st.success(f"Loaded {len(loaded)} routes.")
            st.rerun()
        except Exception as e:
            st.error(f"Failed to load session: {e}")

if not _api_key or not _api_key.startswith("sk-ant"):
    st.warning("Enter your Anthropic API key in the sidebar to continue.")
    st.stop()

try:
    client = anthropic.Anthropic(api_key=_api_key)
except Exception as e:
    st.error(f"Failed to initialize Claude API: {e}")
    st.stop()

if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    pin = st.text_input("Enter access PIN:", type="password")
    if st.button("Authenticate"):
        if pin == "dsny2025":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Invalid PIN.")
    st.stop()

st.success("Authenticated")

# ─── SESSION STATE ──────────────────────────────────────────────────────────────

if 'routes' not in st.session_state:
    st.session_state.routes = []
if 'detail_open' not in st.session_state:
    st.session_state.detail_open = {}

# ─── CLAUDE VISION ─────────────────────────────────────────────────────────────

def compress_image(image_bytes: bytes, max_bytes: int = 4_500_000) -> tuple[bytes, str]:
    from PIL import Image
    if len(image_bytes) <= max_bytes:
        return image_bytes, "image/jpeg"
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    quality = 85
    while quality >= 30:
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality)
        data = buf.getvalue()
        if len(data) <= max_bytes:
            return data, "image/jpeg"
        quality -= 10
    img = img.resize((img.width // 2, img.height // 2), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=75)
    return buf.getvalue(), "image/jpeg"


def process_image_with_claude(image_bytes: bytes, media_type: str) -> Optional[Dict]:
    try:
        image_bytes, media_type = compress_image(image_bytes)
        b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
        prompt = (
            "This is a DSNY DS-659 Route Narrative form. "
            "Extract ALL route information and return ONLY valid JSON.\n\n"
            "JSON structure:\n{\n"
            '  "section": "section code",\n'
            '  "route": "route number",\n'
            '  "district": "district code",\n'
            '  "material": "material description",\n'
            '  "vehicle_type": "vehicle type",\n'
            '  "itsas": [\n    {"number": 1, "street": "STREET NAME", "from_cross": "FROM", "to_cross": "TO", "side": "B"}\n  ],\n'
            '  "extraction_confidence": "high|medium|low"\n}\n\n'
            "Rules:\n- Extract EVERY ITSA row\n- Use UPPERCASE for street names\n- Side: B=Both, R=Right, L=Left\n- Return ONLY the JSON"
        )
        msg = client.messages.create(
            model="claude-opus-4-5-20251101", max_tokens=4096,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": prompt}
            ]}]
        )
        raw = "".join(b.text for b in msg.content if b.type == "text").strip()
        if debug_mode:
            with st.expander("Claude raw response (Debug)"):
                st.text(raw)
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
        return json.loads(raw)
    except json.JSONDecodeError as e:
        st.error(f"Claude returned invalid JSON: {e}")
        return None
    except Exception as e:
        st.error(f"Claude API error: {e}")
        return None


def process_pdf_with_claude(file_bytes: bytes) -> Optional[Dict]:
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        text = "".join(p.extract_text() or "" for p in reader.pages)
        if not text.strip():
            st.warning("PDF has no extractable text — try uploading a photo instead.")
            return None
        prompt = (
            "This is DSNY DS-659 route sheet text. Extract all data and return ONLY valid JSON:\n"
            '{"section":"","route":"","district":"","material":"","itsas":[{"number":1,"street":"","from_cross":"","to_cross":"","side":"B"}],'
            '"extraction_confidence":"high|medium|low"}\n\nText:\n' + text
        )
        msg = client.messages.create(model="claude-opus-4-5-20251101", max_tokens=4096,
                                     messages=[{"role": "user", "content": prompt}])
        raw = "".join(b.text for b in msg.content if b.type == "text").strip()
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
        return json.loads(raw)
    except Exception as e:
        st.error(f"PDF processing error: {e}")
        return None


# ─── WORK LEFT OUT — DS-659 EXCEL ──────────────────────────────────────────────

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "ds659_template.xlsx")

def generate_work_left_out(missed_df: pd.DataFrame, route_info: dict) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws['A3'] = route_info.get('district', '') or ws['A3'].value
    ws['D3'] = route_info.get('section', '') or ws['D3'].value
    ws['H1'] = route_info.get('vehicle_type', '') or ws['H1'].value
    ws['J1'] = route_info.get('material', '') or ws['J1'].value
    for row_num in range(8, 26):
        for col in ['A', 'B', 'C', 'D', 'H', 'J', 'L', 'M', 'N']:
            ws[f'{col}{row_num}'] = None
    for i, (_, r) in enumerate(missed_df.iterrows()):
        row_num = 8 + i
        if row_num > 25:
            break
        ws[f'A{row_num}'] = route_info.get('section', '')
        ws[f'B{row_num}'] = r.get('ITSA #', '')
        ws[f'C{row_num}'] = r.get('Side', 'B')
        ws[f'D{row_num}'] = r.get('Street', '')
        ws[f'H{row_num}'] = r.get('From', '')
        ws[f'J{row_num}'] = r.get('To', '')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_truly_missed_df(r: dict) -> pd.DataFrame:
    """Return SKIPPED rows that are NOT manually overridden."""
    df = r["df"]
    manual_overrides = r.get('manual_overrides', {})
    manually_done_keys = {k for k, v in manual_overrides.items() if v}
    missed_df = df[
        df["Status"].str.contains("SKIPPED") &
        ~df["ITSA #"].astype(str).isin(manually_done_keys)
    ]
    return missed_df


# ─── DS-332 DAILY ROUTE ASSIGNMENT PDF ────────────────────────────────────────

def generate_ds332_pdf(route_entries: list, date_str: str = None, garage: str = '') -> bytes:
    """Generate DS-332 Daily Route Assignment PDF — landscape, matching actual DSNY form."""
    if not date_str:
        date_str = datetime.now().strftime("%m/%d/%Y")

    page = landscape(letter)  # 11 x 8.5 inches
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=0.4*inch, rightMargin=0.4*inch,
        topMargin=0.35*inch, bottomMargin=0.35*inch
    )

    styles = getSampleStyleSheet()
    center_bold = ParagraphStyle('CenterBold', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER)
    center_sm = ParagraphStyle('CenterSm', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)
    left_sm = ParagraphStyle('LeftSm', fontName='Helvetica', fontSize=8, alignment=TA_LEFT)

    elements = []
    W = page[0] - 0.8*inch  # usable width

    # ── Header block ──────────────────────────────────────────────────────────
    first_cj = route_entries[0].get('claude_json', {}) if route_entries else {}
    district  = first_cj.get('district', '')
    section_h = first_cj.get('section', '')

    garage_text = f"   <b>GARAGE:</b> {garage}" if garage else ""
    hdr_data = [
        [
            Paragraph("NEW YORK CITY\nDEPARTMENT OF SANITATION", center_bold),
            Paragraph("DAILY ROUTE ASSIGNMENT\nDS-332", center_bold),
            Paragraph(
                f"<b>DATE:</b> {date_str}          "
                f"<b>DISTRICT:</b> {district}          "
                f"<b>SECTION:</b> {section_h}"
                f"{garage_text}",
                left_sm
            ),
        ]
    ]
    hdr_table = Table(hdr_data, colWidths=[2.6*inch, 3.0*inch, W - 5.6*inch])
    hdr_table.setStyle(TableStyle([
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0,0), (-1, -1), 4),
        ('BOX',         (0, 0), (-1, -1), 1, colors.black),
        ('LINEBEFORE',  (1, 0), (1, -1), 1, colors.black),
        ('LINEBEFORE',  (2, 0), (2, -1), 1, colors.black),
    ]))
    elements.append(hdr_table)
    elements.append(Spacer(1, 0.08*inch))

    # ── Main data table ────────────────────────────────────────────────────────
    # Columns: # | Truck # | Route | Section | District | Material | Sanitation Workers | % Done | ITSAs Done | ITSAs Missed | Remarks
    col_labels = ['#', 'Truck #', 'Route', 'Section', 'District', 'Material',
                  'Sanitation Workers', '% Done', 'Done', 'Missed', 'Remarks']
    col_w = [0.25*inch, 0.75*inch, 0.55*inch, 0.65*inch, 0.65*inch, 0.85*inch,
             2.4*inch, 0.5*inch, 0.45*inch, 0.5*inch, 1.55*inch]

    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=7.5, alignment=TA_CENTER, leading=9)
    cell_left  = ParagraphStyle('CellL', fontName='Helvetica', fontSize=7.5, alignment=TA_LEFT, leading=9)

    table_data = [col_labels]
    for i, r in enumerate(route_entries):
        cj      = r.get('claude_json', {})
        pct     = r.get('pct', 0)
        done    = r.get('done', 0)
        total   = r.get('total', 0)
        missed  = total - done
        workers = r.get('workers', '').strip() or ''
        shift_start = r.get('shift_start', '')
        shift_end   = r.get('shift_end', '')
        notes       = r.get('notes', '')
        manual_overrides = r.get('manual_overrides', {})
        manual_count = sum(1 for v in manual_overrides.values() if v)

        remarks_parts = []
        if shift_start or shift_end:
            remarks_parts.append(f"{shift_start}-{shift_end}")
        if notes:
            remarks_parts.append(notes)
        if manual_count > 0:
            remarks_parts.append(f"({manual_count} manual)")
        remarks = ' '.join(remarks_parts).strip()

        table_data.append([
            str(i + 1),
            r.get('truck', ''),
            r.get('route', ''),
            cj.get('section', ''),
            cj.get('district', ''),
            cj.get('material', ''),
            workers,
            f"{pct}%",
            str(done),
            str(missed),
            remarks,
        ])

    # Pad to at least 20 rows so form looks complete
    while len(table_data) < 21:
        table_data.append(['', '', '', '', '', '', '', '', '', '', ''])

    main_table = Table(table_data, colWidths=col_w, repeatRows=1)
    main_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 7.5),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, 0), 'MIDDLE'),
        # Data rows
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 1), (-1, -1), 'CENTER'),
        ('ALIGN',         (6, 1), (6, -1), 'LEFT'),   # workers left-aligned
        ('ALIGN',         (10, 1),(10, -1),'LEFT'),   # remarks left-aligned
        ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
        # Alternating rows
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        # Grid
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.black),
        # Row heights
        ('ROWHEIGHT',     (0, 0), (0, 0), 16),
        ('ROWHEIGHT',     (0, 1), (-1, -1), 14),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(main_table)
    elements.append(Spacer(1, 0.12*inch))

    # ── Summary row ────────────────────────────────────────────────────────────
    total_done_all = sum(r.get('done', 0) for r in route_entries)
    total_itsas    = sum(r.get('total', 0) for r in route_entries)
    total_missed   = total_itsas - total_done_all
    overall_pct    = round(total_done_all / total_itsas * 100, 1) if total_itsas else 0.0

    summary_data = [[
        Paragraph(f"<b>TOTAL ROUTES:</b> {len(route_entries)}", left_sm),
        Paragraph(f"<b>TOTAL ITSAs:</b> {total_itsas}", left_sm),
        Paragraph(f"<b>COMPLETED:</b> {total_done_all}", left_sm),
        Paragraph(f"<b>MISSED:</b> {total_missed}", left_sm),
        Paragraph(f"<b>OVERALL:</b> {overall_pct}%", left_sm),
    ]]
    summary_table = Table(summary_data, colWidths=[W/5]*5)
    summary_table.setStyle(TableStyle([
        ('BOX',          (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID',    (0, 0), (-1, -1), 0.3, colors.grey),
        ('BACKGROUND',   (0, 0), (-1, -1), colors.HexColor('#e8e8e8')),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.15*inch))

    # ── Signature block ────────────────────────────────────────────────────────
    sig_data = [[
        Paragraph("Supervisor Signature: _______________________________", left_sm),
        Paragraph(f"Date: {date_str}", left_sm),
        Paragraph("Title: Supervisor MTS", left_sm),
        Paragraph("Badge #: 5104", left_sm),
        Paragraph("Time: ____________", left_sm),
    ]]
    sig_table = Table(sig_data, colWidths=[W*0.35, W*0.15, W*0.2, W*0.15, W*0.15])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE',     (0, 0), (-1, -1), 8),
        ('TOPPADDING',   (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
        ('BOX',          (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID',    (0, 0), (-1, -1), 0.3, colors.grey),
        ('LEFTPADDING',  (0, 0), (-1, -1), 4),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


# ─── GPS PARSING ───────────────────────────────────────────────────────────────

def parse_rastrac_csv(gps_df: pd.DataFrame) -> set:
    streets_visited = set()
    addr_col = next((c for c in gps_df.columns if 'addr' in c.lower() or c.lower() == 'address'), None)
    if not addr_col:
        return streets_visited
    for addr in gps_df[addr_col].dropna():
        parts = str(addr).strip().split(',')
        if parts:
            street_only = re.sub(r'^\d+\s+', '', parts[0].strip()).strip().upper()
            if street_only:
                streets_visited.add(street_only)
    return streets_visited


def normalize_street(name: str) -> str:
    name = name.upper().strip()
    for full, abbr in {'AVENUE':'AVE','STREET':'ST','BOULEVARD':'BLVD','DRIVE':'DR','COURT':'CT',
                       'PLACE':'PL','ROAD':'RD','LANE':'LN','TERRACE':'TER','HIGHWAY':'HWY','PARKWAY':'PKWY'}.items():
        name = re.sub(r'\b' + full + r'\b', abbr, name)
    return name.strip()


def verify_itsas_against_gps(itsas: List[Dict], streets_visited: set) -> pd.DataFrame:
    rows = []
    norm_visited = {normalize_street(s) for s in streets_visited}
    for itsa in itsas:
        num = itsa.get('number', '?')
        street = str(itsa.get('street', '')).strip()
        from_cross = itsa.get('from_cross', '')
        to_cross = itsa.get('to_cross', '')
        side = itsa.get('side', 'B')
        norm_street = normalize_street(street)
        matched = norm_street in norm_visited
        if not matched:
            street_words = set(norm_street.split())
            for visited in norm_visited:
                if len(street_words & set(visited.split())) >= min(2, len(street_words)):
                    matched = True
                    break
        status = "✅ DONE" if matched else "❌ SKIPPED"
        rows.append({"ITSA #": num, "Street": street, "From": from_cross, "To": to_cross, "Side": side, "Status": status})
    return pd.DataFrame(rows)


# ─── BOROUGH INFERENCE ─────────────────────────────────────────────────────────

DISTRICT_TO_BOROUGH = {'Q':'Queens, NY','M':'Manhattan, NY','BX':'Bronx, NY','BK':'Brooklyn, NY','SI':'Staten Island, NY'}

def infer_borough(claude_json: dict) -> str:
    district = str(claude_json.get('district', '')).upper().strip()
    section = str(claude_json.get('section', '')).upper().strip()
    for key, borough in DISTRICT_TO_BOROUGH.items():
        if district.startswith(key):
            return borough
    for key, borough in DISTRICT_TO_BOROUGH.items():
        if section.startswith(key):
            return borough
    return 'New York, NY'


def build_maps_url(streets: List[str], borough: str) -> str:
    encoded = "/".join(s.replace(' ', '+') + ',+' + borough.replace(' ', '+').replace(',', '') for s in streets)
    return f"https://www.google.com/maps/dir/My+Location/{encoded}"


def chunk_list(lst: list, n: int) -> List[list]:
    return [lst[i:i+n] for i in range(0, len(lst), n)]


# ─── MANUAL OVERRIDE CALLBACK ──────────────────────────────────────────────────

def on_manual_override_change(route_idx: int, itsa_num):
    key = f"manual_{route_idx}_{itsa_num}"
    new_val = st.session_state.get(key, False)
    route = st.session_state.routes[route_idx]
    overrides = route.get('manual_overrides', {})
    overrides[str(itsa_num)] = new_val
    route['manual_overrides'] = overrides
    # Recalculate
    df = route['df']
    gps_done = len(df[df['Status'].str.contains('DONE')])
    manual_done = sum(1 for v in overrides.values() if v)
    total = route['total']
    done = gps_done + manual_done
    route['done'] = done
    route['pct'] = round(done / total * 100, 1) if total > 0 else 0.0


# ─── UPLOAD PANEL ─────────────────────────────────────────────────────────────

with st.expander("➕ Add a Route", expanded=len(st.session_state.routes) == 0):
    col_truck, col_route = st.columns(2)
    with col_truck:
        input_truck = st.text_input("Truck #", placeholder="e.g. 24DP-421", key="input_truck")
    with col_route:
        input_route = st.text_input("Route #", placeholder="e.g. M4", key="input_route")

    route_file = st.file_uploader("Upload DS-659 route sheet photo or PDF",
                                  type=["jpg", "jpeg", "png", "pdf"], key="upload_route_file")
    gps_file = st.file_uploader("Upload Rastrac GPS CSV", type=["csv"], key="upload_gps_file")
    add_btn = st.button("Add Route", type="primary", key="btn_add_route")

    if add_btn:
        errors = []
        if not input_truck.strip(): errors.append("Truck # is required.")
        if not input_route.strip(): errors.append("Route # is required.")
        if not route_file: errors.append("DS-659 route sheet file is required.")
        if not gps_file: errors.append("GPS CSV file is required.")
        if errors:
            for e in errors:
                st.error(e)
        else:
            with st.spinner(f"Processing Truck {input_truck.strip()} / Route {input_route.strip()}..."):
                file_bytes = route_file.read()
                ext = route_file.name.split('.')[-1].lower()
                if ext == 'pdf':
                    claude_json = process_pdf_with_claude(file_bytes)
                else:
                    media_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png'}
                    claude_json = process_image_with_claude(file_bytes, media_map.get(ext, 'image/jpeg'))

                gps_streets = set()
                try:
                    gps_df = pd.read_csv(gps_file)
                    gps_streets = parse_rastrac_csv(gps_df)
                except Exception as e:
                    st.error(f"Failed to load GPS file: {e}")
                    claude_json = None

                if claude_json and gps_streets is not None:
                    itsas = claude_json.get('itsas', [])
                    if not itsas:
                        st.error("No ITSAs found in route sheet. Cannot add route.")
                    else:
                        df = verify_itsas_against_gps(itsas, gps_streets)
                        total = len(df)
                        done = len(df[df['Status'].str.contains('DONE')])
                        pct = round(done / total * 100, 1) if total > 0 else 0.0
                        route_entry = {
                            "truck": input_truck.strip(),
                            "route": input_route.strip(),
                            "claude_json": claude_json,
                            "gps_streets": gps_streets,
                            "df": df,
                            "done": done,
                            "total": total,
                            "pct": pct,
                            "workers": "",
                            "shift_start": "",
                            "shift_end": "",
                            "notes": "",
                            "manual_overrides": {},
                        }
                        st.session_state.routes.append(route_entry)
                        st.toast(f"✅ Truck {input_truck.strip()} / Route {input_route.strip()} added")
                        st.rerun()
                elif not claude_json:
                    st.error("Failed to parse route sheet. Please try again.")

    # ─── BATCH UPLOAD SECTION ───────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📸 Batch Upload Multiple Routes")

    batch_route_files = st.file_uploader(
        "Upload multiple DS-659 route sheets",
        type=["jpg", "jpeg", "png", "pdf"],
        accept_multiple_files=True,
        key="batch_route_files"
    )
    batch_gps_file = st.file_uploader(
        "Upload Rastrac GPS CSV (shared for all)",
        type=["csv"],
        key="batch_gps_file"
    )
    process_batch_btn = st.button("🚀 Process Batch", type="primary", key="btn_process_batch")

    if process_batch_btn:
        batch_errors = []
        if not batch_route_files:
            batch_errors.append("Please upload at least one route sheet file.")
        if not batch_gps_file:
            batch_errors.append("Please upload a GPS CSV file for the batch.")
        if batch_errors:
            for e in batch_errors:
                st.error(e)
        else:
            # Parse shared GPS once
            shared_gps_streets = set()
            try:
                batch_gps_df = pd.read_csv(batch_gps_file)
                shared_gps_streets = parse_rastrac_csv(batch_gps_df)
            except Exception as e:
                st.error(f"Failed to load GPS file: {e}")
                shared_gps_streets = None

            if shared_gps_streets is not None:
                processed_count = 0
                batch_progress = st.progress(0)
                batch_status = st.empty()

                for i, batch_file in enumerate(batch_route_files):
                    truck_auto = f"TBD-{i + 1}"
                    route_auto = f"BATCH-{i + 1}"
                    batch_status.text(f"Processing {i + 1}/{len(batch_route_files)}: {batch_file.name}...")

                    with st.spinner(f"Processing file {i + 1}/{len(batch_route_files)}: {batch_file.name}"):
                        file_bytes = batch_file.read()
                        ext = batch_file.name.split('.')[-1].lower()
                        if ext == 'pdf':
                            claude_json = process_pdf_with_claude(file_bytes)
                        else:
                            media_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png'}
                            claude_json = process_image_with_claude(file_bytes, media_map.get(ext, 'image/jpeg'))

                        if claude_json:
                            itsas = claude_json.get('itsas', [])
                            if not itsas:
                                st.warning(f"No ITSAs found in {batch_file.name} — skipping.")
                            else:
                                df = verify_itsas_against_gps(itsas, shared_gps_streets)
                                total = len(df)
                                done = len(df[df['Status'].str.contains('DONE')])
                                pct = round(done / total * 100, 1) if total > 0 else 0.0
                                route_entry = {
                                    "truck": truck_auto,
                                    "route": route_auto,
                                    "claude_json": claude_json,
                                    "gps_streets": shared_gps_streets,
                                    "df": df,
                                    "done": done,
                                    "total": total,
                                    "pct": pct,
                                    "workers": "",
                                    "shift_start": "",
                                    "shift_end": "",
                                    "notes": "",
                                    "manual_overrides": {},
                                }
                                st.session_state.routes.append(route_entry)
                                processed_count += 1
                        else:
                            st.warning(f"Failed to parse {batch_file.name} — skipping.")

                    batch_progress.progress((i + 1) / len(batch_route_files))

                batch_status.empty()
                batch_progress.empty()
                if processed_count > 0:
                    st.success(f"Processed {processed_count} route{'s' if processed_count != 1 else ''}")
                    st.rerun()
                else:
                    st.error("No routes were successfully processed.")


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

routes = st.session_state.routes
n_routes = len(routes)

st.header(f"📊 Route Dashboard — {n_routes} route{'s' if n_routes != 1 else ''}")

if n_routes == 0:
    st.info("No routes loaded yet. Use the **➕ Add a Route** panel above to get started.")
else:
    COLS = 3
    for row_start in range(0, n_routes, COLS):
        card_cols = st.columns(COLS)
        for col_idx in range(COLS):
            route_idx = row_start + col_idx
            if route_idx >= n_routes:
                break

            r = routes[route_idx]
            truck = r["truck"]
            route_label = r["route"]
            cj = r["claude_json"]
            done = r["done"]
            total = r["total"]
            pct = r["pct"]
            section = cj.get("section", "?")
            district = cj.get("district", "?")
            missed_count = total - done
            manual_overrides = r.get('manual_overrides', {})
            manual_count = sum(1 for v in manual_overrides.values() if v)

            with card_cols[col_idx]:
                with st.container(border=True):
                    st.markdown(f"### 🚛 {truck} · Route {route_label}")
                    st.markdown(f"**Section:** {section} &nbsp;|&nbsp; **District:** {district}")

                    # Completion threshold alerts
                    if pct < 70:
                        st.error(f"🔴 {pct}% — Needs Attention")
                    elif pct < 85:
                        st.warning(f"🟡 {pct}% — Partial")
                    else:
                        st.success(f"✅ {pct}% — Good")

                    st.progress(pct / 100 if total > 0 else 0)

                    # Show manual count if any
                    if manual_count > 0:
                        st.markdown(f"✅ {done} done ({manual_count} manual) &nbsp; ❌ {missed_count} missed")
                    else:
                        st.markdown(f"✅ {done} done &nbsp; ❌ {missed_count} missed")

                    # Inline truck / route edit
                    edit_truck_col, edit_route_col = st.columns(2)
                    with edit_truck_col:
                        new_truck = st.text_input(
                            "Truck #",
                            value=st.session_state.routes[route_idx].get('truck', ''),
                            key=f"edit_truck_{route_idx}",
                            label_visibility="visible",
                        )
                        if new_truck != st.session_state.routes[route_idx].get('truck', ''):
                            st.session_state.routes[route_idx]['truck'] = new_truck
                    with edit_route_col:
                        new_route = st.text_input(
                            "Route #",
                            value=st.session_state.routes[route_idx].get('route', ''),
                            key=f"edit_route_{route_idx}",
                            label_visibility="visible",
                        )
                        if new_route != st.session_state.routes[route_idx].get('route', ''):
                            st.session_state.routes[route_idx]['route'] = new_route

                    # Shift time fields
                    time_col1, time_col2 = st.columns(2)
                    with time_col1:
                        shift_start_val = st.text_input(
                            "Start Time",
                            value=st.session_state.routes[route_idx].get('shift_start', ''),
                            placeholder="06:00",
                            key=f"shift_start_{route_idx}",
                        )
                        if shift_start_val != st.session_state.routes[route_idx].get('shift_start', ''):
                            st.session_state.routes[route_idx]['shift_start'] = shift_start_val
                    with time_col2:
                        shift_end_val = st.text_input(
                            "End Time",
                            value=st.session_state.routes[route_idx].get('shift_end', ''),
                            placeholder="14:00",
                            key=f"shift_end_{route_idx}",
                        )
                        if shift_end_val != st.session_state.routes[route_idx].get('shift_end', ''):
                            st.session_state.routes[route_idx]['shift_end'] = shift_end_val

                    # Sanitation Workers input
                    workers_val = st.text_input(
                        "👷 Sanitation Workers",
                        value=st.session_state.routes[route_idx].get('workers', ''),
                        placeholder="e.g. Smith J., Jones R.",
                        key=f"workers_{route_idx}",
                    )
                    if workers_val != st.session_state.routes[route_idx].get('workers', ''):
                        st.session_state.routes[route_idx]['workers'] = workers_val

                    # Route notes
                    notes_val = st.text_area(
                        "📝 Notes",
                        value=st.session_state.routes[route_idx].get('notes', ''),
                        placeholder="Road closures, driver issues, etc.",
                        key=f"notes_{route_idx}",
                        height=68,
                    )
                    if notes_val != st.session_state.routes[route_idx].get('notes', ''):
                        st.session_state.routes[route_idx]['notes'] = notes_val

                    btn_col1, btn_col2, btn_col3 = st.columns(3)

                    with btn_col1:
                        toggle_key = f"detail_open_{route_idx}"
                        if toggle_key not in st.session_state.detail_open:
                            st.session_state.detail_open[toggle_key] = False
                        if st.button("Details ▼", key=f"btn_details_{route_idx}"):
                            st.session_state.detail_open[toggle_key] = not st.session_state.detail_open[toggle_key]
                            st.rerun()

                    with btn_col2:
                        missed_df = get_truly_missed_df(r)
                        if not missed_df.empty and os.path.exists(TEMPLATE_PATH):
                            try:
                                wlo_bytes = generate_work_left_out(missed_df, cj)
                                sec = cj.get('section', 'SEC')
                                rte = cj.get('route', 'RTE')
                                st.download_button(
                                    "📋 Work Left Out",
                                    data=wlo_bytes,
                                    file_name=f"Work_Left_Out_{sec}_{rte}_{truck}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"dl_wlo_{route_idx}"
                                )
                            except Exception as e:
                                st.warning(f"WLO error: {e}")
                        else:
                            st.button("📋 Work Left Out", disabled=True, key=f"dl_wlo_disabled_{route_idx}")

                    with btn_col3:
                        if st.button("🗑️ Delete", key=f"btn_delete_{route_idx}", type="secondary"):
                            st.session_state.routes.pop(route_idx)
                            st.session_state.detail_open.pop(f"detail_open_{route_idx}", None)
                            st.rerun()

        # Detail views
        for col_idx in range(COLS):
            route_idx = row_start + col_idx
            if route_idx >= n_routes:
                break
            toggle_key = f"detail_open_{route_idx}"
            if st.session_state.detail_open.get(toggle_key, False):
                r = routes[route_idx]
                truck = r["truck"]
                route_label = r["route"]
                cj = r["claude_json"]
                df = r["df"]
                done = r["done"]
                total = r["total"]
                pct = r["pct"]
                borough = infer_borough(cj)
                manual_overrides = r.get('manual_overrides', {})

                st.markdown(f"---\n#### 🚛 {truck} · Route {route_label} — Detail View")
                tab1, tab2 = st.tabs(["📋 ITSA Breakdown", "🗺️ Navigation"])

                with tab1:
                    # Build display df reflecting current overrides
                    manual_overrides = r.get('manual_overrides', {})
                    display_rows = []
                    for _, row in df.iterrows():
                        itsa_num = str(row['ITSA #'])
                        gps_status = row['Status']
                        if '✅' in gps_status:
                            status = '✅ GPS'
                        elif manual_overrides.get(itsa_num, False):
                            status = '✅ MANUAL'
                        else:
                            status = '❌ SKIPPED'
                        display_rows.append({
                            'ITSA #': row['ITSA #'],
                            'Street': row['Street'],
                            'From': row['From'],
                            'To': row['To'],
                            'Side': row['Side'],
                            'Status': status
                        })
                    display_df = pd.DataFrame(display_rows)
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    # Count for summary
                    manual_count = sum(1 for v in manual_overrides.values() if v)
                    gps_done = len([row for row in display_rows if row['Status'] == '✅ GPS'])
                    total_done = gps_done + manual_count
                    total_rows = len(display_rows)
                    st.markdown(f"**{total_done} of {total_rows} ITSAs completed ({r.get('pct',0)}%)**")

                    # Manual override section — only show skipped/manual rows
                    skipped_rows = [row for row in display_rows if row['Status'] in ('❌ SKIPPED', '✅ MANUAL')]
                    if skipped_rows:
                        st.markdown("---")
                        st.markdown("**✏️ Manual Overrides** *(supervisor verification)*")
                        for row in skipped_rows:
                            itsa_num = str(row['ITSA #'])
                            is_manual = manual_overrides.get(itsa_num, False)
                            new_val = st.checkbox(
                                f"ITSA {row['ITSA #']} — {row['Street']} ({row['From']} → {row['To']})",
                                value=is_manual,
                                key=f"manual_{route_idx}_{row['ITSA #']}"
                            )
                            if new_val != is_manual:
                                st.session_state.routes[route_idx]['manual_overrides'][itsa_num] = new_val
                                # Recalculate
                                new_manual_count = sum(1 for v in st.session_state.routes[route_idx]['manual_overrides'].values() if v)
                                gps_count = len(df[df['Status'].str.contains('DONE')])
                                new_done = gps_count + new_manual_count
                                new_pct = round(new_done / len(df) * 100, 1) if len(df) > 0 else 0.0
                                st.session_state.routes[route_idx]['done'] = new_done
                                st.session_state.routes[route_idx]['pct'] = new_pct
                                st.rerun()

                with tab2:
                    all_streets = df["Street"].tolist()
                    # Use truly missed (not manually overridden) for nav
                    truly_missed_df = get_truly_missed_df(r)
                    missed_streets = truly_missed_df["Street"].tolist()

                    st.subheader("🗺️ Ride Full Route")
                    for chunk_idx, chunk in enumerate(chunk_list(all_streets, 6)):
                        start_itsa = chunk_idx * 6 + 1
                        end_itsa = start_itsa + len(chunk) - 1
                        url = build_maps_url(chunk, borough)
                        st.markdown(f"[Group {chunk_idx + 1} (ITSAs {start_itsa}–{end_itsa}) →]({url})")

                    st.subheader("🔴 Missed Streets Only")
                    if missed_streets:
                        if len(missed_streets) <= 6:
                            url = build_maps_url(missed_streets, borough)
                            st.markdown(f"[Navigate All Missed ({len(missed_streets)} streets) →]({url})")
                        else:
                            for chunk_idx, chunk in enumerate(chunk_list(missed_streets, 6)):
                                url = build_maps_url(chunk, borough)
                                start_n = chunk_idx * 6 + 1
                                end_n = start_n + len(chunk) - 1
                                st.markdown(f"[Missed Group {chunk_idx + 1} (streets {start_n}–{end_n}) →]({url})")
                        st.markdown("**Individual missed ITSAs:**")
                        for _, row in truly_missed_df.iterrows():
                            nav_url = ("https://www.google.com/maps/dir/My+Location/"
                                       + row["Street"].replace(" ", "+")
                                       + ",+" + borough.replace(" ", "+").replace(",", ""))
                            st.markdown(f"ITSA {row['ITSA #']} — {row['Street']} ({row['From']} → {row['To']}) [Navigate]({nav_url})")
                    else:
                        st.success("No missed streets — all ITSAs completed! 🎉")
                st.markdown("---")


# ─── SUMMARY BAR ──────────────────────────────────────────────────────────────

if n_routes > 0:
    total_done = sum(r["done"] for r in routes)
    total_all = sum(r["total"] for r in routes)
    overall_pct = round(total_done / total_all * 100, 1) if total_all > 0 else 0.0

    st.divider()
    st.markdown(f"**Overall: {total_done}/{total_all} ITSAs complete ({overall_pct}%) across {n_routes} route{'s' if n_routes != 1 else ''}**")

    col_zip, col_ds332 = st.columns(2)

    with col_zip:
        if os.path.exists(TEMPLATE_PATH):
            try:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
                    for r in routes:
                        missed_df = get_truly_missed_df(r)
                        if not missed_df.empty:
                            cj = r["claude_json"]
                            wlo_bytes = generate_work_left_out(missed_df, cj)
                            sec = cj.get("section", "SEC")
                            rte = cj.get("route", "RTE")
                            zf.writestr(f"Work_Left_Out_{sec}_{rte}_{r['truck']}.xlsx", wlo_bytes)
                zip_buf.seek(0)
                st.download_button("📥 Download All Work Left Out", data=zip_buf.getvalue(),
                                   file_name="All_Work_Left_Out.zip", mime="application/zip", key="dl_all_wlo_zip")
            except Exception as e:
                st.warning(f"Could not build zip: {e}")

    with col_ds332:
        try:
            shift_date = st.date_input("Shift Date", value=datetime.now().date(), key="ds332_date")
            date_str = shift_date.strftime("%m/%d/%Y")
            today_str = datetime.now().strftime("%Y%m%d")
            ds332_all_bytes = generate_ds332_pdf(routes, date_str=date_str, garage=st.session_state.get('garage', ''))
            st.download_button("📄 DS-332", data=ds332_all_bytes,
                               file_name=f"DS332_All_{today_str}.pdf", mime="application/pdf", key="dl_ds332_all")
        except Exception as e:
            st.warning(f"DS-332 error: {e}")
